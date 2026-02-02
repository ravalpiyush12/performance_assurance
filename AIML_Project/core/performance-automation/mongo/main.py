import pymongo
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime
from typing import List, Dict, Any
import json
import os

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class MongoToExcelExporter:
    """
    Automated solution to export MongoDB collections to Excel files
    """
    
    def __init__(self, config_file: str = 'mongo_config.json'):
        """
        Initialize the exporter with configuration
        
        Args:
            config_file: Path to JSON configuration file
        """
        self.config = self._load_config(config_file)
        self.client = None
        self.db = None
        
    def _load_config(self, config_file: str) -> Dict:
        """Load configuration from JSON file"""
        try:
            with open(config_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.warning(f"Config file {config_file} not found. Using default configuration.")
            return self._get_default_config()
    
    def _get_default_config(self) -> Dict:
        """Return default configuration"""
        return {
            "mongodb": {
                "host": "localhost",
                "port": 27017,
                "database": "your_database",
                "username": "",
                "password": "",
                "auth_source": "admin"
            },
            "collections": [],
            "output": {
                "directory": "exports",
                "filename_prefix": "mongodb_export",
                "add_timestamp": True
            },
            "excel": {
                "header_style": {
                    "bold": True,
                    "bg_color": "366092",
                    "font_color": "FFFFFF"
                },
                "auto_filter": True,
                "freeze_panes": True
            }
        }
    
    def connect(self) -> bool:
        """
        Establish connection to MongoDB
        
        Returns:
            bool: True if connection successful
        """
        try:
            mongo_config = self.config['mongodb']
            
            # Build connection string
            if mongo_config.get('username') and mongo_config.get('password'):
                connection_string = f"mongodb://{mongo_config['username']}:{mongo_config['password']}@{mongo_config['host']}:{mongo_config['port']}/{mongo_config['database']}?authSource={mongo_config.get('auth_source', 'admin')}"
            else:
                connection_string = f"mongodb://{mongo_config['host']}:{mongo_config['port']}"
            
            self.client = MongoClient(connection_string, serverSelectionTimeoutMS=5000)
            
            # Test connection
            self.client.server_info()
            
            self.db = self.client[mongo_config['database']]
            logger.info(f"Successfully connected to MongoDB database: {mongo_config['database']}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to connect to MongoDB: {str(e)}")
            return False
    
    def get_collection_data(self, collection_name: str, query: Dict = None, 
                           projection: Dict = None, limit: int = None) -> List[Dict]:
        """
        Fetch data from MongoDB collection
        
        Args:
            collection_name: Name of the collection
            query: MongoDB query filter
            projection: Fields to include/exclude
            limit: Maximum number of documents to fetch
            
        Returns:
            List of documents
        """
        try:
            collection = self.db[collection_name]
            
            query = query or {}
            cursor = collection.find(query, projection)
            
            if limit:
                cursor = cursor.limit(limit)
            
            documents = list(cursor)
            logger.info(f"Fetched {len(documents)} documents from collection: {collection_name}")
            return documents
            
        except Exception as e:
            logger.error(f"Error fetching data from {collection_name}: {str(e)}")
            return []
    
    def _flatten_document(self, doc: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """
        Flatten nested MongoDB documents
        
        Args:
            doc: Document to flatten
            parent_key: Parent key for nested fields
            sep: Separator for nested keys
            
        Returns:
            Flattened dictionary
        """
        items = []
        for k, v in doc.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            
            if isinstance(v, dict):
                items.extend(self._flatten_document(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert lists to string representation
                items.append((new_key, str(v)))
            else:
                items.append((new_key, v))
        
        return dict(items)
    
    def _apply_header_style(self, worksheet, header_row: int = 1):
        """Apply styling to header row"""
        style_config = self.config['excel']['header_style']
        
        header_fill = PatternFill(
            start_color=style_config['bg_color'],
            end_color=style_config['bg_color'],
            fill_type="solid"
        )
        header_font = Font(
            bold=style_config['bold'],
            color=style_config['font_color']
        )
        
        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_excel(self, data: List[Dict], output_filename: str, 
                     sheet_name: str = 'Data', flatten: bool = True) -> str:
        """
        Create Excel file from MongoDB data
        
        Args:
            data: List of documents
            output_filename: Output Excel filename
            sheet_name: Name of the Excel sheet
            flatten: Whether to flatten nested documents
            
        Returns:
            Path to created Excel file
        """
        try:
            if not data:
                logger.warning("No data to export")
                return None
            
            # Flatten documents if needed
            if flatten:
                flattened_data = [self._flatten_document(doc) for doc in data]
            else:
                flattened_data = data
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Get all unique keys (columns)
            all_keys = set()
            for doc in flattened_data:
                all_keys.update(doc.keys())
            
            # Remove _id if present and add it at the beginning
            headers = []
            if '_id' in all_keys:
                headers.append('_id')
                all_keys.remove('_id')
            headers.extend(sorted(all_keys))
            
            # Write headers
            ws.append(headers)
            self._apply_header_style(ws)
            
            # Write data
            for doc in flattened_data:
                row = [doc.get(key, '') for key in headers]
                ws.append(row)
            
            # Apply Excel formatting
            if self.config['excel']['auto_filter']:
                ws.auto_filter.ref = ws.dimensions
            
            if self.config['excel']['freeze_panes']:
                ws.freeze_panes = 'A2'
            
            self._auto_adjust_columns(ws)
            
            # Ensure output directory exists
            output_dir = self.config['output']['directory']
            os.makedirs(output_dir, exist_ok=True)
            
            # Build full output path
            output_path = os.path.join(output_dir, output_filename)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel file created successfully: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            return None
    
    def export_collections(self, collections: List[str] = None) -> List[str]:
        """
        Export multiple collections to separate Excel files
        
        Args:
            collections: List of collection names to export
            
        Returns:
            List of created file paths
        """
        if not self.db:
            logger.error("Not connected to MongoDB. Call connect() first.")
            return []
        
        collections = collections or self.config.get('collections', [])
        
        if not collections:
            # Export all collections if none specified
            collections = self.db.list_collection_names()
            logger.info(f"No collections specified. Exporting all collections: {collections}")
        
        created_files = []
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S') if self.config['output']['add_timestamp'] else ''
        
        for collection_name in collections:
            try:
                logger.info(f"Processing collection: {collection_name}")
                
                # Fetch data
                data = self.get_collection_data(collection_name)
                
                if not data:
                    logger.warning(f"No data found in collection: {collection_name}")
                    continue
                
                # Generate filename
                prefix = self.config['output']['filename_prefix']
                if timestamp:
                    filename = f"{prefix}_{collection_name}_{timestamp}.xlsx"
                else:
                    filename = f"{prefix}_{collection_name}.xlsx"
                
                # Create Excel
                output_path = self.create_excel(data, filename, sheet_name=collection_name)
                
                if output_path:
                    created_files.append(output_path)
                
            except Exception as e:
                logger.error(f"Error processing collection {collection_name}: {str(e)}")
        
        return created_files
    
    def export_with_aggregation(self, collection_name: str, pipeline: List[Dict], 
                                output_filename: str, sheet_name: str = 'Data') -> str:
        """
        Export aggregated data to Excel
        
        Args:
            collection_name: Name of the collection
            pipeline: MongoDB aggregation pipeline
            output_filename: Output Excel filename
            sheet_name: Name of the Excel sheet
            
        Returns:
            Path to created Excel file
        """
        try:
            collection = self.db[collection_name]
            data = list(collection.aggregate(pipeline))
            
            logger.info(f"Aggregation returned {len(data)} documents")
            
            return self.create_excel(data, output_filename, sheet_name)
            
        except Exception as e:
            logger.error(f"Error in aggregation export: {str(e)}")
            return None
    
    def close(self):
        """Close MongoDB connection"""
        if self.client:
            self.client.close()
            logger.info("MongoDB connection closed")


def main():
    """
    Main execution function
    """
    # Initialize exporter
    exporter = MongoToExcelExporter('mongo_config.json')
    
    # Connect to MongoDB
    if not exporter.connect():
        logger.error("Failed to connect to MongoDB. Exiting.")
        return
    
    try:
        # Example 1: Export specific collections
        created_files = exporter.export_collections(['users', 'orders', 'products'])
        
        # Example 2: Export with custom query
        # data = exporter.get_collection_data(
        #     'orders',
        #     query={'status': 'completed'},
        #     projection={'_id': 1, 'order_id': 1, 'total': 1}
        # )
        # exporter.create_excel(data, 'completed_orders.xlsx', 'Completed Orders')
        
        # Example 3: Export with aggregation
        # pipeline = [
        #     {'$match': {'status': 'active'}},
        #     {'$group': {'_id': '$category', 'count': {'$sum': 1}}},
        #     {'$sort': {'count': -1}}
        # ]
        # exporter.export_with_aggregation('products', pipeline, 'product_summary.xlsx')
        
        logger.info(f"Export completed. Created {len(created_files)} files:")
        for file_path in created_files:
            logger.info(f"  - {file_path}")
            
    finally:
        # Close connection
        exporter.close()


if __name__ == '__main__':
    main()