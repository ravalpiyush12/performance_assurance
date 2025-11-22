# ======================================================
# API SLA MONITOR v29 – FINAL
# CSV: reports/API_Data.csv (never confused)
# Data Source clearly printed
# ART: Regex (Oracle/CSV) | Round-robin (Demo)
# ======================================================
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re
import warnings
import logging
import os
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter

# --------------------------------------------------------------
# Suppress logs
# --------------------------------------------------------------
logging.getLogger('prophet').setLevel(logging.WARNING)
logging.getLogger('cmdstanpy').setLevel(logging.WARNING)
warnings.filterwarnings("ignore") 

# --------------------------------------------------------------
# ML Forecasting (Prophet)
# --------------------------------------------------------------
try:
    from prophet import Prophet
    PROPHET_AVAILABLE = True
    print("Prophet available → using ML forecasting")
except ImportError:
    PROPHET_AVAILABLE = False
    print("Prophet not installed → using linear fallback")
# ================================================================
# CONFIG – DATA SOURCE CONTROL
# ================================================================
SLA_P95_MS = 2000
SLA_FAILURE_PCT = 0.01

# ========================================
# CONTROL DATA SOURCE HERE
# ========================================
# "CSV"     → Use reports/API_Data.csv
# "ORACLE"  → Use Oracle DB (fallback to CSV)
# "DEMO"    → Force generate demo data
# ========================================
DATA_SOURCE = "DEMO"          # ← CHANGE THIS LINE

OUTPUT_DIR = "reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_FILE = f"{OUTPUT_DIR}/API_Data.csv"
EXCEL_FILE = f"{OUTPUT_DIR}/api_sla_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

# --------------------------------------------------------------
# 1. REGEX → ART (ONLY for Oracle/CSV real data)
# --------------------------------------------------------------
ART_REGEX_MAP = [
    (r"payment.?initiation", "Payment Initiation"),
    (r"payment.?management", "Payment Management"),
    (r"platform",            "Platform"),
    (r"entitlement",         "Entitlement"),
    (r"trade",               "Trade"),
    (r"information.?management", "Information Management"),
    (r"liquidity",           "Liquidity"),
    (r"connect.?api",        "Connect API"),
    (r".*",                  "Others")
]

def _assign_art_from_path(path: str) -> str:
    low = re.sub(r'[-_]', '', path.lower())
    for pat, art in ART_REGEX_MAP:
        if re.search(pat, low):
            return art
    return "Others"

# --------------------------------------------------------------
# 2. AUTO‑CREATE DEMO CSV → reports/API_Data.csv (8 ARTs)
# --------------------------------------------------------------
def create_demo_csv():
    if Path(CSV_FILE).exists():
        return

    print(f"Creating DEMO {CSV_FILE} with 120 APIs + 8 ARTs...")

    arts = [
        "Payment Initiation", "Payment Management", "Platform",
        "Entitlement", "Trade", "Information Management",
        "Liquidity", "Connect API"
    ]

    api_to_art = {f"API_{i:03d}": arts[(i-1) % 8] for i in range(1, 121)}
    print("DEMO ART distribution:", Counter(api_to_art.values()))

    np.random.seed(42)
    apis = list(api_to_art.keys())
    today = datetime(2025, 11, 15).date()
    date_range = pd.date_range(start=today - timedelta(days=30), end=today, freq='D').date

    tiers = {
        "stable":     {"apis": apis[0:70],   "base": 600,  "drift": 0,    "fail": 0.001},
        "unstable":   {"apis": apis[70:90],  "base": 1600, "drift": 15,   "fail": 0.006},
        "high_risk":  {"apis": apis[90:95],  "base": 1500, "drift": 100,  "fail": 0.018},
        "med_risk":   {"apis": apis[95:101], "base": 1300, "drift": 58,   "fail": 0.014},
        "low_risk":   {"apis": apis[101:106],"base": 1100, "drift": 38,   "fail": 0.011},
        "critical":   {"apis": apis[106:120],"base": 3000, "drift": 40,   "fail": 0.035}
    }

    data = []
    for tier, cfg in tiers.items():
        for api in cfg["apis"]:
            art = api_to_art[api]
            p95 = cfg["base"]
            for d_idx, date in enumerate(date_range):
                total = np.random.randint(500, 2000)
                failures = int(total * cfg["fail"])
                p95_val = p95 + cfg["drift"] * d_idx + np.random.normal(0, 60)
                p95_val = max(100, min(5000, p95_val))
                data.append([date, api, art, total, failures, round(p95_val, 1)])

    df = pd.DataFrame(data, columns=["date", "api", "art", "total_count", "failures", "p95_rt"])
    df.to_csv(CSV_FILE, index=False)
    print(f"{CSV_FILE} created → DEMO MODE")

# --------------------------------------------------------------
# 3. LOAD DATA – Oracle / CSV
# --------------------------------------------------------------
def load_from_oracle():
    try:
        import oracledb
        conn = oracledb.connect(user="monitor", password="pass123", dsn="localhost:1521/XE")
        query = """
        SELECT TRUNC(log_time) AS log_date,
               api_name AS api,
               art_name AS art,
               COUNT(*) AS total_count,
               SUM(CASE WHEN status='FAIL' THEN 1 ELSE 0 END) AS failures,
               PERCENTILE_CONT(0.95) WITHIN GROUP (ORDER BY response_time_ms) AS p95_rt
        FROM api_performance_logs
        WHERE log_time >= SYSDATE - 30
        GROUP BY TRUNC(log_time), api_name, art_name
        """
        df = pd.read_sql(query, conn)
        conn.close()
        df["date"] = pd.to_datetime(df["log_date"]).dt.date
        return df[["date", "api", "art", "total_count", "failures", "p95_rt"]]
    except Exception as e:
        print(f"Oracle failed: {e}")
        return None

def load_from_csv():
    if not Path(CSV_FILE).exists():
        return None
    df = pd.read_csv(CSV_FILE, parse_dates=["date"])
    df["date"] = df["date"].dt.date
    return df
# --------------------------------------------------------------
# 4. Load + ART Logic + CLEAR DATA SOURCE LABEL
# --------------------------------------------------------------
print(f"Loading from {DATA_SOURCE}...")

df_raw = None
is_demo = False
data_source_label = ""

if DATA_SOURCE == "ORACLE":
    df_raw = load_from_oracle()
    if df_raw is None:
        print("Oracle failed → falling back to CSV")
        df_raw = load_from_csv()
        data_source_label = "CSV DATA"
    else:
        data_source_label = "ORACLE DATA"

elif DATA_SOURCE == "CSV":
    df_raw = load_from_csv()
    if df_raw is None:
        print("CSV not found → generating DEMO data")
        create_demo_csv()
        df_raw = load_from_csv()
        is_demo = True
        data_source_label = "DEMO DATA"
    else:
        data_source_label = "CSV DATA"

elif DATA_SOURCE == "DEMO":
    print("FORCE DEMO MODE → generating fresh data")
    if Path(CSV_FILE).exists():
        print(f"Removing old {CSV_FILE}...")
        os.remove(CSV_FILE)
    create_demo_csv()
    df_raw = load_from_csv()
    is_demo = True
    data_source_label = "DEMO DATA"

else:
    raise ValueError(f"Invalid DATA_SOURCE: {DATA_SOURCE}. Use 'CSV', 'ORACLE', or 'DEMO'")

# --------------------------------------------------------------
# 5. ART ASSIGNMENT (SOURCE-SPECIFIC)
# --------------------------------------------------------------
if is_demo:
    print("DEMO MODE → Using pre-assigned ARTs (round-robin)")
else:
    if "art" not in df_raw.columns or df_raw["art"].isna().all():
        print("REAL DATA → ART column missing → deriving from API path (regex)")
        df_raw["art"] = df_raw["api"].apply(_assign_art_from_path)
    else:
        blanks = df_raw["art"].isna() | (df_raw["art"].str.strip() == "")
        if blanks.any():
            print(f"REAL DATA → Filling {blanks.sum()} missing ARTs with regex")
            df_raw.loc[blanks, "art"] = df_raw.loc[blanks, "api"].apply(_assign_art_from_path)

# --------------------------------------------------------------
# 6. MANDATORY ART VALIDATION
# --------------------------------------------------------------
still_missing = df_raw[df_raw["art"].isna() | (df_raw["art"].str.strip() == "")]
if not still_missing.empty:
    bad = still_missing["api"].unique().tolist()
    raise ValueError(
        f"ART is MANDATORY. {len(bad)} API(s) still have no ART:\n"
        + ", ".join(bad[:20]) + (" ..." if len(bad) > 20 else "")
    )

apis = sorted(df_raw["api"].unique().tolist())
arts = sorted(df_raw["art"].unique().tolist())
today = datetime.now().date()
yesterday = today - timedelta(days=1)
date_range = pd.date_range(start=today - timedelta(days=30), end=today, freq='D').date

# --------------------------------------------------------------
# 7. PRINT DATA SOURCE CLEARLY
# --------------------------------------------------------------
print("\n" + "="*70)
print(f"DATA SOURCE: {data_source_label}")
print(f"CSV FILE: {CSV_FILE}")
print(f"RUN TIME: {datetime.now().strftime('%d %b %Y, %I:%M %p IST')}")
print("="*70)
print(f"Loaded {len(df_raw)} rows | {len(apis)} APIs | {len(arts)} ARTs")
print("="*70)

# ================================================================
# [Rest of your script: grid, analysis, Excel, plot...]
# ================================================================

# ================================================================
# 8. BUILD FULL GRID (no duplicate‑index errors)
# ================================================================
df_clean = df_raw.drop_duplicates(subset=["api", "date"])

full_grid = pd.DataFrame([
    (api, date, df_raw[df_raw["api"] == api]["art"].iloc[0])
    for api in apis for date in date_range
], columns=["api", "date", "art"])

df_filled = full_grid.merge(df_clean, on=["api", "date", "art"], how="left")
df_filled["p95_rt"] = df_filled.groupby("api")["p95_rt"].ffill()
df_filled[["total_count", "failures"]] = df_filled[["total_count", "failures"]].fillna(0)
df_filled["failure_rate"] = np.where(
    df_filled["total_count"] > 0,
    (df_filled["failures"] / df_filled["total_count"]) * 100,
    0
).round(4)

# ================================================================
# 9. ANALYSIS
# ================================================================
stability = df_filled.groupby("api").agg(
    p95_cv=("p95_rt", lambda x: x.std() / x.mean() if x.mean() > 0 else 0)
).round(3)

summary_data = []
for api in apis:
    api_df = df_filled[df_filled["api"] == api].sort_values("date")
    art = api_df["art"].iloc[0]

    today_row = api_df[api_df["date"] == today]
    yest_row  = api_df[api_df["date"] == yesterday]

    today_p95 = today_row["p95_rt"].iloc[0] if not today_row.empty else None
    today_fr  = today_row["failure_rate"].iloc[0] if not today_row.empty else None
    yest_p95  = yest_row["p95_rt"].iloc[0] if not yest_row.empty else None
    yest_fr   = yest_row["failure_rate"].iloc[0] if not yest_row.empty else None
    best_p95  = api_df["p95_rt"].min()
    best_fr   = api_df["failure_rate"].min()

    cv = stability.loc[api, "p95_cv"]
    stability_label = "Unstable" if cv > 0.25 else "Stable"

    p95_str = f"{today_p95:.0f}ms" if today_p95 else "N/A"
    p95_cmp = (
        f"vs {yest_p95:.0f}ms | {best_p95:.0f}ms best | SLA 2000ms"
        if yest_p95 is not None else f"N/A | {best_p95:.0f}ms best | SLA 2000ms"
    )
    fr_str = f"{today_fr:.3f}%" if today_fr is not None else "N/A"
    fr_cmp = (
        f"vs {yest_fr:.3f}% | {best_fr:.3f}% best | SLA 0.01%"
        if yest_fr is not None else f"N/A | {best_fr:.3f}% best | SLA 0.01%"
    )

    risk = "None"
    if today_p95 and today_p95 > SLA_P95_MS:
        risk = "Critical"
        p95_str += " [BREACHED]"
    else:
        if PROPHET_AVAILABLE and len(api_df) >= 5:
            recent = api_df[api_df["total_count"] > 0][["date", "p95_rt"]].rename(
                columns={"date": "ds", "p95_rt": "y"})
            if len(recent) >= 2:
                m = Prophet(yearly_seasonality=False,
                            weekly_seasonality=True,
                            daily_seasonality=False,
                            interval_width=0.95)
                m.fit(recent)
                future = m.make_future_dataframe(periods=30)
                forecast = m.predict(future)
                future_rt = forecast[forecast["ds"] > pd.Timestamp(today)]
                breach = future_rt[future_rt["yhat_upper"] > SLA_P95_MS]
                if not breach.empty:
                    days = (breach.iloc[0]["ds"].date() - today).days
                    p95_str += f" [Warning] {days}d"
                    risk = "High" if days <= 7 else "Medium" if days <= 14 else "Low"

    summary_data.append({
        "API": api,
        "ART": art,
        "P95": p95_str,
        "P95_Compare": p95_cmp,
        "Fail": fr_str,
        "Fail_Compare": fr_cmp,
        "Best P95": f"{best_p95:.0f}ms",
        "Best Fail": f"{best_fr:.3f}%",
        "Stability": stability_label,
        "Risk": risk,
        "Today_P95": today_p95
    })

summary_df = pd.DataFrame(summary_data)

# ================================================================
# 10. GROUPED DATAFRAMES
# ================================================================
stable_df_full   = summary_df[summary_df["Stability"] == "Stable"].copy()
unstable_df_full = summary_df[summary_df["Stability"] == "Unstable"].copy()
critical_df_full = summary_df[summary_df["Risk"] == "Critical"].copy()
high_risk_df_full = summary_df[summary_df["Risk"] == "High"].copy()
med_risk_df_full  = summary_df[summary_df["Risk"] == "Medium"].copy()
low_risk_df_full  = summary_df[summary_df["Risk"] == "Low"].copy()
no_risk_df_full   = summary_df[summary_df["Risk"] == "None"].copy()

stable_df_full   = stable_df_full.sort_values("Today_P95")
unstable_df_full = unstable_df_full.sort_values("Today_P95", ascending=False)
critical_df_full = critical_df_full.sort_values("Today_P95", ascending=False)
high_risk_df_full = high_risk_df_full.sort_values("Today_P95", ascending=False)
med_risk_df_full  = med_risk_df_full.sort_values("Today_P95", ascending=False)
low_risk_df_full  = low_risk_df_full.sort_values("Today_P95", ascending=False)
no_risk_df_full   = no_risk_df_full.sort_values("Today_P95")

# ================================================================
# 11. TERMINAL SUMMARY + DEFINITIONS
# ================================================================
run_time = datetime.now().strftime("%d %b %Y, %I:%M %p IST")
total_apis = len(apis)

counts = {
    "stable": len(stable_df_full),
    "unstable": len(unstable_df_full),
    "critical": len(critical_df_full),
    "high": len(high_risk_df_full),
    "medium": len(med_risk_df_full),
    "low": len(low_risk_df_full),
    "none": len(no_risk_df_full)
}

summary_text = (
    f"SUMMARY: {counts['stable']} Stable | {counts['unstable']} Unstable | "
    f"{counts['critical']} Critical | {counts['high']} High Risk | {counts['medium']} Medium | "
    f"{counts['low']} Low | {counts['none']} No Risk | Total: {total_apis} APIs"
)

definitions = {
    "Stable": "DEFINITION: CV = std(P95)/mean(P95) ≤ 0.25 → Low volatility in last 30 days",
    "Unstable": "DEFINITION: CV = std(P95)/mean(P95) > 0.25 → High fluctuation in P95",
    "Critical": "DEFINITION: Today P95 > 2000ms → Already breached SLA",
    "High Risk": "DEFINITION: Forecast breach in ≤7 days",
    "Medium Risk": "DEFINITION: Forecast breach in 8–14 days",
    "Low Risk": "DEFINITION: Forecast breach in 15–30 days",
    "No Risk": "DEFINITION: No breach forecasted in next 30 days"
}

print("\n" + "="*120)
print(summary_text)
print("="*120)
for k, v in definitions.items():
    print(f"{k.upper():12}: {v}")
print("="*120)

# ================================================================
# 12. ADD HEADER ROWS (Summary + Definition)
# ================================================================
def add_header_rows(df, def_text):
    if df.empty:
        return pd.DataFrame([{c: "" for c in EXPORT_COLS}])
    h1 = pd.DataFrame([{**{"API": summary_text}, **{c: "" for c in df.columns if c != "API"}}])
    h2 = pd.DataFrame([{**{"API": def_text},    **{c: "" for c in df.columns if c != "API"}}])
    return pd.concat([h1, h2, df.reset_index(drop=True)], ignore_index=True)

EXPORT_COLS = ["API", "ART", "P95", "P95_Compare", "Fail", "Fail_Compare",
               "Best P95", "Best Fail"]

stable_df_full   = add_header_rows(stable_df_full,   definitions["Stable"])
unstable_df_full = add_header_rows(unstable_df_full, definitions["Unstable"])
critical_df_full = add_header_rows(critical_df_full, definitions["Critical"])
high_risk_df_full = add_header_rows(high_risk_df_full, definitions["High Risk"])
med_risk_df_full  = add_header_rows(med_risk_df_full,  definitions["Medium Risk"])
low_risk_df_full  = add_header_rows(low_risk_df_full,  definitions["Low Risk"])
no_risk_df_full   = add_header_rows(no_risk_df_full,   definitions["No Risk"])

# ================================================================
# 13. EXPORT TO EXCEL (ordered sheets)
# ================================================================
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    ordered = [
        ("Stable", stable_df_full),
        ("Unstable", unstable_df_full),
        ("Critical", critical_df_full),
        ("High Risk", high_risk_df_full),
        ("Medium Risk", med_risk_df_full),
        ("Low Risk", low_risk_df_full),
        ("No Risk", no_risk_df_full)
    ]
    for name, df in ordered:
        df_export = df[EXPORT_COLS]
        df_export.to_excel(writer, sheet_name=name, index=False, startrow=2)
        ws = writer.sheets[name]

        # Row 3 = Summary (blue), Row 4 = Definition (yellow)
        for col in range(1, len(EXPORT_COLS) + 1):
            ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=3, column=col).fill = PatternFill(start_color="1F4E79",
                                                         end_color="1F4E79",
                                                         fill_type="solid")
            ws.cell(row=4, column=col).font = Font(bold=True, color="000000")
            ws.cell(row=4, column=col).fill = PatternFill(start_color="FFFF99",
                                                         end_color="FFFF99",
                                                         fill_type="solid")

        # Auto‑width
        for i, col in enumerate(EXPORT_COLS, 1):
            max_len = max(df_export[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_len, 60)

print(f"Excel report → {EXCEL_FILE}")

# ================================================================
# 14. PLOT: Top‑5 worst APIs per ART (merged)
# ================================================================
top5_per_art = (
    summary_df
    .sort_values("Today_P95", ascending=False)
    .groupby("ART")
    .head(5)
    .groupby("ART")
    .apply(lambda g: g.sort_values("Today_P95", ascending=False)["API"].tolist())
)

top_arts = top5_per_art.index[:5].tolist()          # limit to 5 ARTs for the picture

fig = plt.figure(figsize=(30, 12))
gs = fig.add_gridspec(1, 5, hspace=0.3, wspace=0.3)

for idx, art in enumerate(top_arts):
    ax = fig.add_subplot(gs[0, idx])
    apis_in_art = top5_per_art[art]

    for api in apis_in_art:
        sub = df_filled[df_filled["api"] == api].set_index("date")["p95_rt"]
        sub = sub.reindex(date_range)
        ax.plot(sub.index, sub.values,
                label=f"{api} ({sub.values[-1]:.0f}ms)", linewidth=1.8)

    ax.axhline(SLA_P95_MS, color="black", linestyle="--", linewidth=1.5,
               label="SLA 2000ms")
    ax.set_title(f"{art}\nTop 5 APIs", fontsize=12, fontweight='bold')
    ax.set_ylabel("P95 (ms)")
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    ax.grid(alpha=0.3)
    ax.legend(fontsize=8, loc='upper left')

fig.suptitle(f"Top 5 Worst APIs per ART – {run_time}", fontsize=16, y=0.98)
jpeg_file = f"{OUTPUT_DIR}/top5_per_art_{datetime.now().strftime('%Y%m%d_%H%M')}.jpg"
plt.savefig(jpeg_file, dpi=200, bbox_inches='tight', facecolor='white')
plt.close()
print(f"JPEG (Top 5 per ART) → {jpeg_file}")

print(f"Run completed: {run_time}")
